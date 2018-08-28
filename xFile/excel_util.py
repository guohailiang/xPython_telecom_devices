#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/3/4 上午7:27
# @Author  : zhuhao
# @Project : xPython
# @File    : excel_util.py
# @Contact : hyzhuhao891407@163.com 
# @Software : PyCharm
# @Desc: excel操作帮助类
'''Excel文件三个对象
workbook： 工作簿，一个excel文件包含多个sheet。
sheet：工作表，一个workbook有多个，表名识别，如“sheet1”,“sheet2”等。
cell： 单元格，存储数据对象'''

import os
from openpyxl import load_workbook
from openpyxl.styles import Color, Font, Alignment
from openpyxl.styles.colors import BLUE, RED, GREEN, YELLOW
import file_util

class ExcelHelper:
    '''
    xlsx 文件操作类
    '''
    def __init__(self,filename,sheetname=None):
        if not os.path.exists(filename):
            raise Exception('file {0} is not exists.'.format(filename))
        ext=file_util.get_file_ext(filename)
        if not ext in['.xlsx']:
            raise Exception('file extension {0} is not correct.'.format(ext))
        self.filename=filename
        self.wb=load_workbook(filename)
        if sheetname is None:
            self.sheet=self.wb.active
        else:
            self.sheet=self.wb[sheetname]

    def is_open(self):
        '''
        判断excel是否已经打开
        :return:
        '''
        if self.wb:
            return True
        else:
            return False

    def get_sheetnames(self):
        '''
        获取所有的sheet names
        :return:
        '''
        return self.wb.sheetnames

    def get_rownum(self):
        '''
        获取对应sheet的总行数
        :return:
        '''
        rows=self.sheet.max_row
        return rows

    def get_column_num(self):
        '''
        获取对应sheet的总列数
        :return:
        '''
        columns=self.sheet.max_column
        return columns

    def get_cell_value(self,row,col):
        '''
        获取单元格的值
        :param row: 行标
        :param col: 列标
        :return:
        '''
        return self.sheet.cell(row,col).value

    def set_cell_value(self,row,col,value):
        '''
        设置单元格的值
        :param row: 行标
        :param col: 列标
        :param value: 设置的新值
        :return:
        '''
        self.sheet.cell(row, col).value=value

    def cellstyle(self, coord, font, align):
        '''
        设置单元格格式
        font = Font(name=u'宋体', size=14, color=RED, bold=True)
        align = Alignment(horizontal='center', vertical='center')
        使用：ws.cellstyle('B2', font, align)
        :param coord: A1
        :param font: 字体
        :param align: 对齐格式
        :return:
        '''
        cell = self.sheet.cell(coord)
        cell.font = font
        cell.alignment = align

    def save(self):
        '''
        保存数据到打开的excel文件中
        :return:
        '''
        self.wb.save(self.filename)

    def close(self):
        '''关闭打开的excel'''
        if self.is_open():
            self.wb.close()

if __name__=='__main__':
    filename=os.path.join(os.getcwd(),'test.xlsx')
    excel=ExcelHelper(filename)
    excel.set_cell_value(3,1,10)
    print(excel.get_cell_value(3,1))